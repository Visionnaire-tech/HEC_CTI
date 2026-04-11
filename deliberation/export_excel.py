from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from evaluations.models import Grade


def export_deliberation_excel(students):
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliberation"

    # 🔥 HEADER
    ws['A1'] = "GRILLE DE DELIBERATION"
    ws['A1'].font = Font(bold=True, size=14)

    row = 3

    # 🔥 Colonnes fixes
    ws.cell(row=row, column=1, value="Matricule")
    ws.cell(row=row, column=2, value="Nom")

    col = 3

    # 🔥 récupérer structure UE
    first_student = students.first()
    grades = Grade.objects.filter(student=first_student).select_related('course__ue')

    ues = {}

    for g in grades:
        ue = g.course.ue
        if ue.id not in ues:
            ues[ue.id] = []
        ues[ue.id].append(g.course)

    # 🔥 HEADER dynamique
    for ue_courses in ues.values():
        for course in ue_courses:
            ws.cell(row=row, column=col, value=course.code)
            col += 1

        ws.cell(row=row, column=col, value="TNP")
        col += 1

        ws.cell(row=row, column=col, value="MOY UE")
        col += 1

    ws.cell(row=row, column=col, value="MOY GEN")

    # 🔥 DATA
    row += 1

    for student in students:
        col = 1

        ws.cell(row=row, column=col, value=student.matricule)
        col += 1

        ws.cell(row=row, column=col, value=student.nom)
        col += 1

        total_points = 0
        total_credit = 0

        for ue_courses in ues.values():
            ue_tnp = 0
            ue_credit = 0

            for course in ue_courses:
                grade = Grade.objects.filter(
                    student=student,
                    course=course
                ).first()

                note = grade.note_finale if grade else 0

                ws.cell(row=row, column=col, value=note)
                col += 1

                tnp = note * course.credit
                ue_tnp += tnp
                ue_credit += course.credit

            ws.cell(row=row, column=col, value=ue_tnp)
            col += 1

            moyenne = ue_tnp / ue_credit if ue_credit else 0
            ws.cell(row=row, column=col, value=round(moyenne, 2))
            col += 1

            total_points += ue_tnp
            total_credit += ue_credit

        moyenne_gen = total_points / total_credit if total_credit else 0
        ws.cell(row=row, column=col, value=round(moyenne_gen, 2))

        row += 1

    return wb